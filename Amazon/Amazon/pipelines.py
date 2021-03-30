# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import json
import os
import openpyxl
from Amazon.settings import PATH as path

class AmazonJsonPipeline:
    def process_item(self, item, spider):
        content = dict(item)
        content_json = json.dumps(content, ensure_ascii=False) + ',\n'
        with open('{}\\亚马逊.json'.format(path), 'ab') as f:
            f.write(content_json.encode('utf-8'))
        return item

class AmazonExcelPipeline:
    def __init__(self):
        self.filepath = '{}\\亚马逊.xlsx'.format(path)   #文件路径
        # 文件是否存在
        if os.path.exists(self.filepath):
            pass
        else:
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            self.sheet.title = '首页'
            self.wb.save(self.filepath)
        # self.wb = openpyxl.load_workbook(self.filepath)
        self.i = 2
        global i

    def process_item(self, item, spider):
        self.wb = openpyxl.load_workbook(self.filepath)
        self.sheet = self.wb['首页']
        # 首行是否存在
        if self.sheet['A1'].value == None:
            self.sheet['A1'] = '商品名'
            self.sheet['B1'] = '价格'
            self.sheet['C1'] = '图片链接'
            self.sheet['D1'] = '商品链接'
            self.sheet['E1'] = '商品信息'
        else:
            pass
        # 存入数据
        self.sheet['A' + str(self.i)] = item['title']
        self.sheet['B' + str(self.i)] = item['price']
        self.sheet['C' + str(self.i)] = item['img_url']
        self.sheet['D' + str(self.i)] = item['src']
        self.sheet['E' + str(self.i)] = item['info']
        self.i += 1
        self.wb.save(self.filepath)     #保存excel。若运行途中被中断程序，则不会执行close()函数，excel就不会被保存，数据就不会存储进去
        # print('{}保存'.format(self.i))
        return item

    # def close(self):
        # print('close保存')
        # self.wb.save(self.filepath)